import json
import re
import io
import requests
import pdfplumber
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
import streamlit as st
from groq import Groq

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="Datasheet Extractor", page_icon="🔌", layout="wide")
st.title("🔌 Datasheet Thermal Extractor")
st.caption("Upload PDFs or search by part number to extract thermal parameters automatically.")

# ── API Key ───────────────────────────────────────────────────────────────────
client = Groq(api_key=st.secrets["GROQ_API_KEY"])

# ── System Prompt ─────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert electronics engineer extracting data from component datasheets.

Extract the following fields and return ONLY a valid JSON object, no explanation:
- part_number: The primary part number or component name
- package: Package type (e.g. TO-220, QFN, DPAK, SMD, etc.)
- rth_ja: Thermal resistance junction-to-ambient in °C/W (numeric only, no units)
- rth_jc: Thermal resistance junction-to-case in °C/W (numeric only, no units)
- rth_jb: Thermal resistance junction-to-board in °C/W (numeric only, no units)
- tj_max: Maximum junction temperature in °C (numeric only)
- power_dissipation: Maximum power dissipation in W (numeric only, no units)
- confidence: dict with keys rth_ja, rth_jc, rth_jb, tj_max, power_dissipation. Each value must be exactly "high", "low", or "not_found". Never null.
- flags: list of plain-English warnings. Return [] if none.
- source_quote: dict with keys rth_ja, rth_jc, rth_jb, tj_max, power_dissipation. Copy the exact sentence or table row the value came from. Use "not found" if absent.

Rules:
- If a value is not found, return null for the value, "not_found" for confidence, "not found" for source_quote
- confidence and source_quote must always have all 5 keys listed above
- For confidence: "high" = explicit table value, "low" = inferred or found in running text
- Return only the JSON object, nothing else
"""

# ── PDF Helpers ───────────────────────────────────────────────────────────────
def extract_text_from_pdf(file):
    text = ""
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_thermal_section(full_text, window=6000):
    thermal_keywords = [
        r'thermal resistance', r'RthJC', r'RthJA', r'RthJB',
        r'θJA', r'θJC', r'θJB', r'ΘJA', r'ΘJC',
        r'Theta.*JA', r'Theta.*JC',
        r'junction.{0,10}ambient', r'junction.{0,10}case',
        r'maximum ratings', r'absolute maximum',
        r'thermal data', r'thermal information', r'package.*thermal',
        r'Zth', r'thermal impedance', r'ψJA', r'ψJC',
    ]
    pattern = '|'.join(thermal_keywords)
    match = re.search(pattern, full_text, re.IGNORECASE)
    if match:
        start = max(0, match.start() - 200)
        end = min(len(full_text), match.start() + window)
        return full_text[start:end]
    return full_text[:6000]


# ── Source Quote Finder ───────────────────────────────────────────────────────
def find_source_quote(text, value):
    if value is None:
        return "not found in datasheet"
    search_val = str(value).split(".")[0]
    lines = re.split(r'[\n]', text)
    for line in lines:
        if search_val in line and len(line.strip()) > 5:
            return line.strip()[:200]
    return "not found in datasheet"


# ── Normalize LLM output ──────────────────────────────────────────────────────
def normalize_result(data, thermal_text=""):
    conf = data.get("confidence")
    if not isinstance(conf, dict):
        conf = {}
    src = data.get("source_quote")
    if not isinstance(src, dict):
        src = {}
    flags = data.get("flags")
    if not isinstance(flags, list):
        flags = [str(flags)] if flags else []

    thermal_fields = ["rth_ja", "rth_jc", "rth_jb", "tj_max", "power_dissipation"]
    for field in thermal_fields:
        if not conf.get(field):
            conf[field] = "low" if data.get(field) is not None else "not_found"
        src[field] = find_source_quote(thermal_text, data.get(field))

    data["confidence"] = conf
    data["source_quote"] = src
    data["flags"] = flags
    return data


# ── Datasheet Web Search ──────────────────────────────────────────────────────
def search_datasheet_url(part_number):
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    query = f"{part_number} datasheet filetype:pdf"
    ddg_url = f"https://html.duckduckgo.com/html/?q={requests.utils.quote(query)}"
    try:
        resp = requests.get(ddg_url, headers=headers, timeout=10)
        pdf_links = re.findall(r'https?://[^\s"<>&]+\.pdf', resp.text)
        # Prefer manufacturer domains
        preferred = ["st.com", "ti.com", "infineon.com", "nxp.com", "onsemi.com",
                     "rohm.com", "diodes.com", "vishay.com", "alldatasheet.com"]
        for link in pdf_links:
            if any(domain in link for domain in preferred):
                return link
        # Return first PDF found if no preferred domain matched
        return pdf_links[0] if pdf_links else None
    except Exception:
        return None


def download_pdf(url):
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        content_type = resp.headers.get("Content-Type", "")
        if resp.status_code == 200 and ("pdf" in content_type.lower() or url.endswith(".pdf")):
            return io.BytesIO(resp.content)
    except Exception:
        pass
    return None


# ── Extraction ────────────────────────────────────────────────────────────────
def extract_component_data(file, filename):
    full_text = extract_text_from_pdf(file)
    thermal_text = extract_thermal_section(full_text)

    try:
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": f"Extract component data from this datasheet section:\n\n{thermal_text}"}
            ],
            response_format={"type": "json_object"},
            temperature=0,
        )
        raw_output = response.choices[0].message.content.strip()
        json_match = re.search(r'\{.*\}', raw_output, re.DOTALL)
        data = json.loads(json_match.group()) if json_match else json.loads(raw_output)
        data["source_file"] = filename
        data = normalize_result(data, thermal_text)
        return data, None

    except json.JSONDecodeError as e:
        return None, f"JSON parse error: {e}"
    except Exception as e:
        return None, str(e)


# ── Excel Builder ─────────────────────────────────────────────────────────────
def build_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Component Data"

    headers = [
        "Source File", "Part Number", "Package",
        "RthJA (°C/W)", "RthJA Confidence", "RthJA Source",
        "RthJC (°C/W)", "RthJC Confidence", "RthJC Source",
        "RthJB (°C/W)", "RthJB Confidence", "RthJB Source",
        "Tj Max (°C)", "Tj Confidence", "Tj Source",
        "Power Dissipation (W)", "Power Confidence", "Power Source",
        "Flags"
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, r in enumerate(results, 2):
        conf = r.get("confidence", {})
        src = r.get("source_quote", {})
        flags = r.get("flags", [])

        row_data = [
            r.get("source_file"),
            r.get("part_number"),
            r.get("package"),
            r.get("rth_ja"),            conf.get("rth_ja"),            src.get("rth_ja"),
            r.get("rth_jc"),            conf.get("rth_jc"),            src.get("rth_jc"),
            r.get("rth_jb"),            conf.get("rth_jb"),            src.get("rth_jb"),
            r.get("tj_max"),            conf.get("tj_max"),            src.get("tj_max"),
            r.get("power_dissipation"), conf.get("power_dissipation"), src.get("power_dissipation"),
            " | ".join(flags) if flags else ""
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if value == "low":
                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            elif value == "not_found":
                cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── UI ────────────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📎 Upload PDF", "🔍 Search by Part Number"])

files_to_process = []

with tab1:
    uploaded_files = st.file_uploader(
        "Upload Datasheet PDFs",
        type=["pdf"],
        accept_multiple_files=True
    )
    if uploaded_files:
        files_to_process = [(f, f.name) for f in uploaded_files]
        st.write(f"**{len(files_to_process)} file(s) uploaded.**")

with tab2:
    part_input = st.text_area(
        "Enter part numbers (one per line)",
        placeholder="e.g.\nIRF540N\nLM317\nSTM32F103C8T6"
    )
    search_btn = st.button("🔍 Find Datasheets", type="primary")

    if search_btn and part_input.strip():
        parts = [p.strip() for p in part_input.strip().splitlines() if p.strip()]
        for part in parts:
            with st.spinner(f"Searching datasheet for {part}..."):
                pdf_url = search_datasheet_url(part)
                if pdf_url:
                    st.caption(f"✅ Found: {pdf_url}")
                    pdf_file = download_pdf(pdf_url)
                    if pdf_file:
                        files_to_process.append((pdf_file, f"{part}.pdf"))
                    else:
                        st.warning(f"⚠️ Found URL but could not download PDF for {part}")
                else:
                    st.warning(f"⚠️ No datasheet found for {part}")

# ── Process & Extract ─────────────────────────────────────────────────────────
if files_to_process:
    if st.button("🚀 Extract Data", type="primary"):
        results = []
        progress = st.progress(0)
        status = st.empty()

        for i, (file, filename) in enumerate(files_to_process):
            status.write(f"⏳ Processing **{filename}**...")
            data, error = extract_component_data(file, filename)

            if data:
                results.append(data)
            else:
                st.warning(f"⚠️ Failed on {filename}: {error}")
                results.append(normalize_result({
                    "source_file": filename,
                    "part_number": "EXTRACTION FAILED",
                    "package": None, "rth_ja": None, "rth_jc": None,
                    "rth_jb": None, "tj_max": None, "power_dissipation": None,
                    "confidence": {}, "flags": [error], "source_quote": {}
                }))

            progress.progress((i + 1) / len(files_to_process))

        status.success(f"✅ Done! Processed {len(results)} file(s).")

        st.subheader("📊 Extracted Data")
        display_rows = []
        for r in results:
            conf = r.get("confidence", {})
            flags = r.get("flags", [])
            display_rows.append({
                "File": r.get("source_file"),
                "Part No.": r.get("part_number"),
                "Package": r.get("package"),
                "RthJA": r.get("rth_ja"),
                "RthJA ✓": conf.get("rth_ja", "—"),
                "RthJC": r.get("rth_jc"),
                "RthJC ✓": conf.get("rth_jc", "—"),
                "RthJB": r.get("rth_jb"),
                "Tj Max": r.get("tj_max"),
                "Ptot": r.get("power_dissipation"),
                "⚠️ Flags": " | ".join(flags) if flags else "—"
            })

        df = pd.DataFrame(display_rows)
        st.dataframe(df, use_container_width=True)

        all_flags = [(r.get("source_file"), f) for r in results for f in r.get("flags", [])]
        if all_flags:
            st.subheader("⚠️ Flags to Review")
            for source, flag in all_flags:
                st.warning(f"**{source}**: {flag}")

        excel_buffer = build_excel(results)
        st.download_button(
            label="📥 Download Excel",
            data=excel_buffer,
            file_name="extracted_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

else:
    st.info("👆 Upload PDFs or search by part number to get started.")